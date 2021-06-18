import pandas as pd
import os
import datetime
from openpyxl import load_workbook

PROJ = 'C19'
DISEASE = 'COVID-19'

EXCEL_PATH = './DataSheet_20210616.xlsx'
MRN_CTDATE_PATH = './COVID_CTs_20210513CarissaWalter_0608jc_20210616.xlsx'
VIDA_PATH_PREFIX = ''
DCM_PATH_PREFIX = ''

parsed_data = pd.read_excel(EXCEL_PATH, sheet_name=0)
vida_dashboard_data = pd.read_excel(EXCEL_PATH, sheet_name=1)
mrn_ctdate_data = pd.read_excel(MRN_CTDATE_PATH, header=9)

start_vida_case_number = 1381
vida_dashboard_data = vida_dashboard_data[vida_dashboard_data['Case ID']
                                          >= start_vida_case_number]
# subj_dict = {}
# for _, row in mrn_ctdate_data.iterrows():
#     mrn = row['mrn']
#     ctdate = row['date'].strftime('%Y%m%d') if type(
#         row['date']) == datetime.datetime else ''
#     subj_id = row['Subj']
#     img_id = str(row['Time'])
#     subj = subj_id + '_' + 'IN' + img_id
#     subj_dict[subj] = [mrn, ctdate]

subj_ctdate_dict = {}
for _, row in mrn_ctdate_data.iterrows():
    mrn = row['mrn']
    ctdate = row['date'].strftime('%Y%m%d') if type(
        row['date']) == pd.Timestamp else ''
    subj_id = row['Subj']
    img_id = str(row['Time'])
    subj = subj_id + '_' + ctdate
    subj_ctdate_dict[subj] = [mrn, img_id]


def construct_new_rows(vida_dashboard_data):
    last_row_parsed = parsed_data.iloc[-1]
    last_row_raw = vida_dashboard_data.iloc[-1]
    start_index = last_row_parsed['VidaCaseID'] + 1
    end_index = last_row_raw['Case ID'] + 1

    rows_to_append = []
    for i in range(start_index, end_index):
        row = {}
        row_data_raw = vida_dashboard_data.loc[vida_dashboard_data['Case ID'] == i]

        if row_data_raw.empty:
            print(f'Case ID {i} info does not exist')
            continue

        #accession_number = row_data_raw['Accession Number'].tolist()[0]
        #subj_id = accession_number.split('_')[0]
        #img_id = accession_number.split('_')[1]
        subj_id = row_data_raw['Patient ID'].tolist()[0].split('_')[0]
        ctdate = row_data_raw['Acquisition Date'].tolist()[
            0].strftime('%Y%m%d')
        try:
            img_id = 'IN' + subj_ctdate_dict[subj_id + '_' + ctdate][1]
            mrn = subj_ctdate_dict[subj_id + '_' + ctdate][0]
        except:
            print(f'{subj_id} + {ctdate} cant be found')

        row['Proj'] = PROJ
        row['Subj'] = subj_id
        row['VidaCaseID'] = i
        row['VidaBy'] = ''
        # row['MRN'] = subj_dict[subj_id + '_' + img_id][0]
        row['MRN'] = mrn
        row['Vida Progress'] = row_data_raw['Process status'].tolist()[0]
        row['Progress'] = 'DL Segmentation Done'
        # row['ScanDate'] = subj_dict[subj_id + '_' + img_id][1]
        row['ScanDate'] = ctdate
        row['IN/EX'] = img_id
        row['CT Protocol'] = row_data_raw['Series Desc'].tolist()[0]
        row['Disease'] = DISEASE
        row['SliceThickness_mm'] = row_data_raw['Slice Thickness'].tolist()[0]
        row['ScannerVender'] = row_data_raw['Scanner'].tolist()[0]
        row['ScannerModel'] = row_data_raw['Scanner Model'].tolist()[0]
        row['Kernel'] = row_data_raw['Kernel'].tolist()[0]
        row['Comments'] = ''
        row['VIDA path'] = VIDA_PATH_PREFIX + str(i)
        row['Report path'] = DCM_PATH_PREFIX

        # if row['Vida Status'] == 'Success':
        # rows_to_append.append(row)
        rows_to_append.append(row)

    return pd.DataFrame(rows_to_append)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


df = construct_new_rows(vida_dashboard_data)
append_df_to_excel(EXCEL_PATH, df, sheet_name='Sheet1',
                   header=None, index=False)
