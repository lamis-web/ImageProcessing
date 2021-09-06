# Usage
# - python update_datasheet.py {PROJ}
# Dependency
# - VidaDatasheet.xlsx
import os
import argparse
import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook


XLSX_PATH = r"E:\common\Taewon\oneDrive\OneDrive - University of Kansas Medical Center\VidaDataSheet.xlsx"

# Create an argumnet parser
# parser = argparse.ArgumentParser(
#     description='Update VidaDataSheet.xlsx')
# parser.add_argument('proj', metavar='proj', type=str,
#                     help='Project Name')
# args = parser.parse_args()

# PROJ = args.proj


def _find_first_case_id_to_update(vida_datasheet_df: DataFrame) -> int:
    return int(vida_datasheet_df.iloc[-1]["VidaCaseID"]) + 1


def _constuct_df_to_append(
    vida_start_case_id: int, vida_dashboard_df: DataFrame
) -> DataFrame:
    vida_dashboard_df_filtered = vida_dashboard_df[
        vida_dashboard_df["Case ID"] > vida_start_case_id
    ]
    rows_to_append = []
    for _, df_row in vida_dashboard_df_filtered.iterrows():
        row = {}
        row["Proj"] = ""
        row["Subj"] = df_row["Patient Name"]
        row["VidaCaseID"] = df_row["Case ID"]
        row["VidaBy"] = ""
        row["MRN"] = ""
        row["Vida Progress"] = df_row["Process status"]
        row["Progress"] = ""
        row["ScanDate"] = int(df_row["Acquisition Date"].strftime("%Y%m%d"))
        row["IN/EX"] = df_row["Scan Type"]
        row["CT Protocol"] = df_row["Series Desc"]
        row["Disease"] = ""
        row["SliceThickness_mm"] = df_row["Slice Thickness"]
        row["ScannerVender"] = df_row["Scanner"]
        row["ScannerModel"] = df_row["Scanner Model"]
        row["Kernel"] = df_row["Kernel"]
        row["Comments"] = ""
        row["VIDA path"] = ""
        row["Report path"] = ""

        rows_to_append.append(row)

    return pd.DataFrame(rows_to_append)


def _append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs
        )
        return

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="replace")

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    print(startrow)

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def update_vida_datasheet(vida_datasheet_df: DataFrame, vida_dashboard_df: DataFrame):
    vida_start_case_id = _find_first_case_id_to_update(vida_datasheet_df)
    df_to_append = _constuct_df_to_append(vida_start_case_id, vida_dashboard_df)
    # print(df_to_append)
    _append_df_to_excel(XLSX_PATH, df_to_append, sheet_name="Sheet1", header=None, index=False)



if __name__ == "__main__":
    vida_datasheet_df = pd.read_excel(XLSX_PATH, sheet_name=0)
    vida_dashboard_df = pd.read_excel(XLSX_PATH, sheet_name=1)
    update_vida_datasheet(vida_datasheet_df, vida_dashboard_df)
